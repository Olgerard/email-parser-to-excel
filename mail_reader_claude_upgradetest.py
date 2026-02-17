import base64
import os
import sys
import tempfile
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from PyPDF2 import PdfReader
from anthropic import Anthropic
import json
from openpyxl import load_workbook
from dotenv import load_dotenv
import tkinter
import customtkinter
import threading
from tkinter import filedialog
from tkinter import ttk
import tkinter.font as tkFont
from openpyxl.styles import numbers
from tkcalendar import DateEntry
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

load_dotenv()

client = Anthropic(api_key=os.getenv("claude_api_key"))

if not client:
    print("❌ Vul eerst je .env bestand in!")
    quit()

def email_to_text_gmail(service, date, label):
    query = f'after:{datetime.strptime(date, "%d-%b-%Y").strftime("%Y/%m/%d")}'
    print(query)
    if label:
        all_labels = service.users().labels().list(userId='me').execute().get('labels', [])
        label_id = next((l['id'] for l in all_labels if l['name'] == label), None)
    else:
        label_id = None
        finish_label.configure(text="Geen map ingevuld!", text_color="red")
        quit()

    results = service.users().messages().list(
        userId='me',
        q=query,
        maxResults=1000,
        labelIds=[label_id] if label_id else None
    ).execute()

    messages = results.get('messages', [])
    messages.reverse()
    extracted_mails = []

    for msg in messages:
        m = service.users().messages().get(userId='me', id=msg['id']).execute()
        mail_timestamp = int(m.get('internalDate', 0)) / 1000
        mail_date_str = datetime.fromtimestamp(mail_timestamp).strftime("%d-%m-%Y")
        payload = m['payload']
        parts = payload.get('parts', [])
        complete_mail = f"Verzenddatum: {mail_date_str} - "

        def walk_parts(parts_list):
            nonlocal complete_mail
            for part in parts_list:
                mime_type = part.get('mimeType', '')
                body_data = part.get('body', {}).get('data')
                if body_data:
                    decoded = base64.urlsafe_b64decode(body_data).decode(errors="ignore")
                    if mime_type == 'text/plain':
                        complete_mail += decoded
                    elif mime_type == 'application/pdf':
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                            tmp_file.write(base64.urlsafe_b64decode(body_data))
                            tmp_filepath = tmp_file.name
                        try:
                            reader = PdfReader(tmp_filepath)
                            pdf_text = "\n".join([page.extract_text() for page in reader.pages])
                            complete_mail += "Text from PDF:" + pdf_text
                        finally:
                            os.remove(tmp_filepath)
                if 'parts' in part:
                    walk_parts(part['parts'])

        walk_parts(parts)
        extracted_mails.append(complete_mail)

    return extracted_mails

def extract_flight_data(email_text):
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=8000,
        temperature=0,
        messages=[
            {"role": "user", "content": f"""Analyseer deze email (bijlagen starten met 'Text from PDF:') en extraheer reis-/boekingsgegevens als JSON.
    
    Formatten per type:
    
    [{{
        "type": "vlucht of trein/bus",
        "boekingsdatum": "dd/mm/jjjj (meestal de datum waarop de mail gestuurd is)",
        "datums": [
            {{"datum": "dd/mm/jjjj (datum heenreis)"}},
            {{"datum": "dd/mm/jjjj (datum terugreis, indien aanwezig)"}}
        ],
        "passagiers": [
            {{"naam": "voornaam achternaam"}}
        ],
        "bestemming": [
            {{"vlucht": "Stad van vertrek - Stad van aankomst"}},
            {{"vlucht": "Terugvlucht (indien aanwezig)"}}
        ],
        "prijs": "totale eindprijs (bijv. 123.45)",
        "PNR": "boekingscode",
        "airline": "naam van de luchtvaartmaatschappij"
    }}]
    
    Hotels, als het leeg is moet het op dezelfde manier als bij vluchten:
    [
    {{
        "type": "hotel",
        "boekingsdatum": "",
        "datum": "Incheck datum (dd/mm) - uitcheck datum (dd/mm)",
        "passagiers": [
            {{"naam": "voornaam achternaam"}}
        ],
        "bestemming": "Naam hotel, Stad",
        "prijs": "",
        "PNR": "",
        "airline": ""
    }}]
    
    Deze vorm voor refunds, als het leeg is moet het op dezelfde manier als bij vluchten:
    [{{
        "type": "refund",
        "boekingsdatum": "dd/mm/jjjj",
        "datum": "dd/mm/jjjj",
        "passagiers": [
            {{"naam": "voornaam achternaam"}}
        ],
        "bestemming": "Enkel heenvlucht (vb Brussel - Amsterdam)",
        "prijs": vb -123.45",
        "PNR": "",
        "airline": ""
    }}]
    
    Regels:
    **Namen & Titels:**
    - Verwijder titels: Mr, Mrs, Ms → niet in naam
    - Format: "Voornaam Achternaam" (hoofdletters aan begin, NOOIT drukletters)
    - LOT format "ACHTERNAAM VOORNAAM Mr" → draai om naar "Voornaam Achternaam"
    
    **Plaatsnamen:**
    - Altijd Nederlands en voluit (vertaal indien nodig)
    - Alleen stadsnaam, geen luchthavennaam
    
    **Vluchten:**
    - Tussenstops samenvoegen: Amsterdam - Warschau - Wroclaw → Amsterdam - Wroclaw
    
    **Maatschappijen:**
    - Verkort: "LOT Airlines" → "LOT", "KLM Airlines" → "KLM", "TAP Air Portugal" → "TAP", "Expedia TAAP" → "Expedia", "Booking.com" → "Booking"
    - Altijd hoofdletter eerste letter
    - NMBS: PNR = DNR van ticket
    
    **Prijzen:**
    - Euro: alleen cijfer (123.45)
    - Andere valuta: cijfer + code (179.99 PLN)
    - Expedia: neem bedrag bij "Betaald aan Expedia", anders hoogste bedrag
    
    **Hotels:**
    - Passagier ontbreekt EN "Company" of "Pieter Smit" staat vermeld → passagier: "Company of Pieter Smit [BE/NL] [aantal]x"
    
    **Refunds (KLM):**
    - Vaak alleen: boekingsdatum, PNR, mogelijk naam
    - Rest velden leeg laten
    
    **Lege velden:**
    - Leeg string "", NOOIT "N/A" of null
    
    **Output:**
    - Alleen pure JSON, geen ```json tags (Dit zorgt ervoor dat heel het programma crasht en is extreem belangrijk!!!), geen tekst eromheen
    - Exact formaat zoals voorbeelden
    EMAIL:
    \"\"\"
    {email_text}
    \"\"\"
    """}
        ]
    )
    return response.content[0].text

def extracted_data_to_excel(ws, row, item):
    passagiers = item.get("passagiers", [])
    first_passenger = True
    current_row = row

    for passenger in passagiers:
        ws.insert_rows(current_row)
        if first_passenger:
            try:
                date_obj = datetime.strptime(item["boekingsdatum"], "%d/%m/%Y").date()
                cell = ws.cell(row=current_row, column=1, value=date_obj)
                cell.number_format = "DD/MM/YYYY"
            except (ValueError, KeyError):
                ws.cell(row=current_row, column=1).value = item.get("boekingsdatum", "")
            try:
                date_obj = datetime.strptime(item["datum"], "%d/%m/%Y").date()
                cell = ws.cell(row=current_row, column=2, value=date_obj)
                cell.number_format = "DD/MM/YYYY"
            except (ValueError, KeyError):
                ws.cell(row=current_row, column=2).value = item.get("datum", "")
            ws.cell(row=current_row, column=3).value = ""
            ws.cell(row=current_row, column=5).value = item.get("bestemming", "")
            ws.cell(row=current_row, column=6).value = ""
            try:
                price = float(item["prijs"])
                cell = ws.cell(row=current_row, column=11, value=price)
                cell.number_format = "#,##0.00"
            except (ValueError, KeyError):
                ws.cell(row=current_row, column=11).value = item.get("prijs", "")
            ws.cell(row=current_row, column=12).value = item.get("PNR", "")
            ws.cell(row=current_row, column=13).value = item.get("airline", "")
        ws.cell(row=current_row, column=4).value = passenger.get("naam", "")
        current_row += 1
        first_passenger = False

def extracted_flightdata_to_excel(ws, row, item):
    first_destination = True
    bestemmingen = item.get("bestemming",[])
    datums = item.get("datums", [])
    passagiers = item.get("passagiers", [])

    current_row = row

    for dest_index, dest in enumerate(bestemmingen):
        first_passenger = True

        datum = datums[dest_index].get("datum", "") if dest_index < len(datums) else ""

        for passenger in passagiers:

            ws.insert_rows(current_row)

            if first_passenger and first_destination:
                try:
                    date_obj = datetime.strptime(item.get("boekingsdatum", ""), "%d/%m/%Y").date()
                    cell = ws.cell(row=current_row, column=1, value=date_obj)
                    cell.number_format = "DD/MM/YYYY"
                except (ValueError, AttributeError):
                    ws.cell(row=current_row, column=1).value = item.get("boekingsdatum", "")

                try:
                    date_obj = datetime.strptime(datum, "%d/%m/%Y").date()
                    cell = ws.cell(row=current_row, column=2, value=date_obj)
                    cell.number_format = "DD/MM/YYYY"
                except (ValueError, AttributeError):
                    ws.cell(row=current_row, column=2).value = datum

                ws.cell(row=current_row, column=3).value = ""
                ws.cell(row=current_row, column=4).value = passenger.get("naam", "")
                ws.cell(row=current_row, column=5).value = dest.get("vlucht", "")
                ws.cell(row=current_row, column=6).value = ""

                try:
                    price = float(item["prijs"])
                    cell = ws.cell(row=current_row, column=11, value=price)
                    cell.number_format = "#,##0.00"
                except (ValueError, KeyError):
                    ws.cell(row=current_row, column=11).value = item.get("prijs", "")

                ws.cell(row=current_row, column=12).value = item.get("PNR", "")
                ws.cell(row=current_row, column=13).value = item.get("airline", "")

            else:
                try:
                    date_obj = datetime.strptime(datum, "%d/%m/%Y").date()
                    cell = ws.cell(row=current_row, column=2, value=date_obj)
                    cell.number_format = "DD/MM/YYYY"
                except (ValueError, KeyError):
                    ws.cell(row=current_row, column=2).value = datum

                ws.cell(row=current_row, column=4).value = passenger.get("naam", "")
                ws.cell(row=current_row, column=5).value = dest.get("vlucht", "")

            current_row += 1
            first_passenger = False

        first_destination = False

def initialize_excel_sheet(excel_path, map):
    wb = load_workbook(excel_path)
    prefix_map = {
        "INBOX/Dossiers/0 Excel NL": "NL",
        "INBOX/Dossiers/0 Excel BE": "BE",
        "INBOX/Dossiers/0 Inv DE": "DE",
        "INBOX/Dossiers/0 Inv FR": "FR",
        "INBOX/Dossiers/0 Inv Nightliner": "Nightliner"
    }
    prefix = prefix_map.get(map, "")
    timestamp = datetime.now().strftime('%d-%m _ %H-%M')
    title = f"{prefix} - {timestamp}" if prefix else timestamp
    ws = wb.create_sheet(title=title)

    ws.cell(row=1, column=1).value = "Boekingsdatum"
    ws.cell(row=1, column=2).value = "Datum"
    ws.cell(row=1, column=3).value = "Tour"
    ws.cell(row=1, column=4).value = "Passagier"
    ws.cell(row=1, column=5).value = "Bestemming"
    ws.cell(row=1, column=6).value = "Prijs"
    ws.cell(row=1, column=7).value = "Fee"
    ws.cell(row=1, column=8).value = ""
    ws.cell(row=1, column=9).value = "Voorgesteld alternatief"
    ws.cell(row=1, column=10).value = "Missed/ Earned Savings"
    ws.cell(row=1, column=11).value = "Prijs Excel"
    ws.cell(row=1, column=12).value = "PNR"
    ws.cell(row=1, column=13).value = "Airline"

    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    row = 2
    ws.cell(row=row, column=1).value = "Vluchten"
    ws.cell(row=row, column=1).fill = highlight
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1).value = "Trein/bus"
    ws.cell(row=row, column=1).fill = highlight
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1).value = "Hotels"
    ws.cell(row=row, column=1).fill = highlight
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1).value = "Refunds"
    ws.cell(row=row, column=1).fill = highlight
    ws.cell(row=row, column=1).font = Font(bold=True)

    wb.save(excel_path)
    return title

def append_item_to_excel(item, excel_path, sheet_name):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    ticket_type = item.get("type","")

    section_names = {
        "vlucht": "Vluchten",
        "trein/bus": "Trein/bus",
        "hotel": "Hotels"
    }
    section_name = section_names.get(ticket_type, "Refunds")

    section_start = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == section_name:
            section_start = row
            break

    if section_start == 0:
        print(f"⚠️ WAARSCHUWING: Sectie '{section_name}' niet gevonden! Item overgeslagen.")
        return

    next_section_row = None
    for row in range(section_start + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value in ["Vluchten", "Trein/bus", "Hotels", "Refunds"]:
            next_section_row = row
            break

    if next_section_row:
        insert_row = next_section_row - 1
    else:
        insert_row = ws.max_row + 1

    if ticket_type == "vlucht" or ticket_type == "trein/bus":
        extracted_flightdata_to_excel(ws, insert_row, item)
    else:
        extracted_data_to_excel(ws, insert_row, item)

    format_excel_cells(ws)
    wb.save(excel_path)

def format_excel_cells(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = numbers.FORMAT_DATE_DDMMYY

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.number_format = "#,##0.00"

def main(service, date, map, excel):
    finish_label.configure(text="Bezig met mails lezen en in Excel zetten!", text_color="white")
    progress_label.configure(text= "")

    try:
        sheet_name = initialize_excel_sheet(excel, map)
    except Exception as e:
        finish_label.configure(text=f"Error bij aanmaken Excel sheet: {e}", text_color="red")
        print(e)
        return

    email_list = email_to_text_gmail(service, date, map)
    number_of_mails = len(email_list)
    number_of_handled_mails = 0
    progress_label.configure(text=f"0 van de {number_of_mails} uitgelezen")

    for m in email_list:
        try:
            json_string = extract_flight_data(m)
            print(json_string)
            parsed = json.loads(json_string)
            if isinstance(parsed, list):
                if len(parsed) > 0:
                    item = parsed[0]
                else:
                    print("⚠️ Lege lijst ontvangen van AI, email overgeslagen")
                    continue
            else:
                item = parsed

            append_item_to_excel(item, excel, sheet_name)

            number_of_handled_mails += 1
            progress_label.configure(text=f"{number_of_handled_mails} van de {number_of_mails} uitgelezen")

        except json.JSONDecodeError as e:
            finish_label.configure(text="AI geeft onleesbare vorm terug", text_color="red")
            print("JSON kon niet gelezen worden:", e)
            print(json_string)
            continue

        except Exception as e:
            print(f"Onverwachte error: {e}")
            continue

    finish_label.configure(text="Klaar! Selecteer een andere map of sluit het programma.", text_color="green")

def start_main_thread():
    thread = threading.Thread(target=lambda: main(service, date_entry.get_date().strftime("%d-%b-%Y"), map.get(), excel_path.get()))
    thread.start()

def browse_file():
    filepath = filedialog.askopenfilename(
        title="Selecteer Excel-bestand",
        filetypes=[("Excel bestanden", "*.xlsx")],
    )
    if filepath:
        excel_path.set(filepath)
        file_label.configure(text=f"Geselecteerd: {os.path.basename(filepath)}", text_color="green")
        run_btn.configure(state=tkinter.NORMAL)


def get_gmail_service():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        try:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open('token.json', 'w') as token:
                token.write(creds.to_json())
        except Exception as e:
            if os.path.exists('token.json'):
                os.remove('token.json')
            return get_gmail_service()
    return build('gmail', 'v1', credentials=creds)

def logout():
    if os.path.exists("token.json"):
        os.remove("token.json")
        app.destroy()
        sys.exit()
    print("Uitgelogd: token verwijderd.")


service = get_gmail_service()

# labels ophalen
results = service.users().labels().list(userId='me').execute()
labels = results.get('labels', [])
mailbox_names = [label['name'] for label in labels]

# System setting
customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")

# App frame
app = customtkinter.CTk()
app.geometry("720x480")
app.title("Excel assistent")

# UI Elements
date_title = customtkinter.CTkLabel(app, text="Vul de datum van de eerste mail in")
date_title.pack(pady=(10,0))

date = tkinter.StringVar()
date_entry = DateEntry(app, date_pattern='dd/mm/yyyy')
date_entry.configure(font=tkFont.Font(size=14))
date_entry.pack(padx=10, pady=(10,20))

map_title = customtkinter.CTkLabel(app, text="Vul de naam van de map in (vb. NL)")
map_title.pack()

all_mailboxes = mailbox_names.copy()
map = tkinter.StringVar()
cb = ttk.Combobox(app, values=all_mailboxes, textvariable=map, width=25, height=40)
cb.configure(font=tkFont.Font(size=14) )
cb.pack(pady=(0,10))

excel_path = tkinter.StringVar()
browse_btn = customtkinter.CTkButton(app, text="Kies Excel-bestand", command=browse_file)
browse_btn.pack(pady=(20,0))

file_label = customtkinter.CTkLabel(app, text="Geen bestand geselecteerd", text_color="red")
file_label.pack()

run_btn = customtkinter.CTkButton(app, text="Start", command=start_main_thread)
run_btn.pack(pady=(20,10))
run_btn.configure(state=tkinter.DISABLED)

logout_btn = customtkinter.CTkButton(app, text="Log uit", command=logout)
logout_btn.pack(pady=(20,10))

finish_label = customtkinter.CTkLabel(app, text = "")
finish_label.pack()

progress_label = customtkinter.CTkLabel(app, text = "")
progress_label.pack()

def filter_mailboxes(event):
    typed = map.get().lower()

    if typed == "":
        cb["values"] = all_mailboxes
    else:
        cb["values"] = [
            m for m in all_mailboxes if typed in m.lower()
        ]

cb.bind("<KeyRelease>", filter_mailboxes)

# Run app
app.mainloop()


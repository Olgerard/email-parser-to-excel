import json
import sys
import tempfile
import threading
import pdfplumber
from datetime import datetime
import html2text
from anthropic import Anthropic
import requests
from tkcalendar import DateEntry
import tkinter.font as tkFont
import tkinter
from tkinter import ttk, filedialog
import customtkinter
import os
import msal
from dotenv import load_dotenv
import re
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.styles import Font, PatternFill
import base64
import webbrowser
# Outlook login data
load_dotenv()
CLIENT_ID = os.getenv("APPLICATION_ID")
USERNAME = os.getenv("OUTLOOK_USER")
CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "token_cache.bin")
SCOPES = ["Mail.Read"]
EXTRACTION_SCHEMA = {
    "type": "array",
    "items": {
        "anyOf": [
            {
                "type": "object",
                "properties": {
                    "type": {"type": "string", "enum": ["vlucht", "trein/bus"]},
                    "boekingsdatum": {"type": "string"},
                    "datums": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {"datum": {"type": "string"}},
                            "required": ["datum"],
                            "additionalProperties": False
                        }
                    },
                    "passagiers": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {"naam": {"type": "string"}},
                            "required": ["naam"],
                            "additionalProperties": False
                        }
                    },
                    "bestemming": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {"vlucht": {"type": "string"}},
                            "required": ["vlucht"],
                            "additionalProperties": False
                        }
                    },
                    "prijs": {"type": "string"},
                    "PNR": {"type": "string"},
                    "airline": {"type": "string"}
                },
                "required": ["type", "boekingsdatum", "datums", "passagiers", "bestemming", "prijs", "PNR", "airline"],
                "additionalProperties": False
            },
            {
                "type": "object",
                "properties": {
                    "type": {"type": "string", "enum": ["hotel", "refund"]},
                    "boekingsdatum": {"type": "string"},
                    "datum": {"type": "string"},
                    "passagiers": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {"naam": {"type": "string"}},
                            "required": ["naam"],
                            "additionalProperties": False
                        }
                    },
                    "bestemming": {"type": "string"},
                    "prijs": {"type": "string"},
                    "PNR": {"type": "string"},
                    "airline": {"type": "string"}
                },
                "required": ["type", "boekingsdatum", "datum", "passagiers", "bestemming", "prijs", "PNR", "airline"],
                "additionalProperties": False
            }
        ]
    }
}

client = Anthropic(api_key=os.getenv("claude_api_key"))

# Get token to access Outlook mail
def get_token(device_flow_callback, verbose=False):
    # Cache to store token
    cache = msal.SerializableTokenCache()
    if os.path.exists(CACHE_FILE):
        cache.deserialize(open(CACHE_FILE, "r").read())

    # Create outlook object to make connection
    app = msal.PublicClientApplication(
        CLIENT_ID,
        authority="https://login.microsoftonline.com/common",
        token_cache=cache,
    )

    # See if an account is logged in locally, try to get access token with refresh token or with logged in account
    accounts = app.get_accounts(username=USERNAME)
    result = app.acquire_token_silent(SCOPES, account=accounts[0]) if accounts else None

    # Manual login if needed
    if not result:
        if verbose:
            print("Geen geldig token in cache, eenmalige login nodig")

        flow = app.initiate_device_flow(scopes=SCOPES)
        device_flow_callback(flow)
        if "user_code" not in flow:
            raise Exception(f"Device flow mislukt: {flow.get('error')}: {flow.get('error_description', flow)}")

        print(flow["message"])
        result = app.acquire_token_by_device_flow(flow)

    # Save cache
    if cache.has_state_changed:
        with open(CACHE_FILE, "w") as f:
            f.write(cache.serialize())

    if "access_token" not in result:
        raise RuntimeError(result.get("error_description", "Kon geen token krijgen"))

    return result["access_token"]

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
def graph_headers(token):
    return {"Authorization": f"Bearer {token}"}

def clean_text(msg):
    msg = re.sub(r'\n\s*\n+', '\n', msg) #1 enter in plaats van meerdere
    msg = re.sub(r'[ \t]+', ' ', msg) #spatie voor tab
    return msg.strip()

def html_to_text(html):
    converter = html2text.HTML2Text()
    converter.ignore_links = True
    converter.ignore_images = True
    converter.body_width = 0
    return converter.handle(html)

def get_pdf_attachments_text(token, message_id):
    headers = graph_headers(token)
    url = f"{GRAPH_BASE}/me/messages/{message_id}/attachments"
    response = requests.get(url, headers=headers)
    response.raise_for_status()

    text = ""
    for attachment in response.json().get("value", []):
        if attachment.get("contentType") == "application/pdf" and "contentBytes" in attachment:
            pdf_bytes = base64.b64decode(attachment["contentBytes"])
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                tmp_file.write(pdf_bytes)
                tmp_filepath = tmp_file.name
            try:
                with pdfplumber.open(tmp_filepath) as pdf:
                    pdf_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
                text += "Text from PDF:" + pdf_text
            finally:
                os.remove(tmp_filepath)
    return text

def message_to_text(token, message):
    received = message.get("receivedDateTime", "")
    mail_date = ""
    if received:
        mail_date = datetime.fromisoformat(received.replace("Z", "+00:00")).strftime("%d-%m-%Y")

    body = message.get("body", {})
    content = body.get("content", "")
    body_text = html_to_text(content) if body.get("contentType") == "html" else content

    complete_mail = f"Verzenddatum: {mail_date} - {body_text}"

    if message.get("hasAttachments"):
        complete_mail += get_pdf_attachments_text(token, message["id"])

    return clean_text(complete_mail)

# Fetch emails and turn into strings
def get_all_email_data(token, folder_id, since_date_iso):
    headers = graph_headers(token)
    url = (f"{GRAPH_BASE}/me/mailFolders/{folder_id}/messages"
           f"?$filter=receivedDateTime ge {since_date_iso}"
           f"&$select=id,receivedDateTime,body,hasAttachments"
           f"&$top=50&$orderby=receivedDateTime asc")

    messages_text = []
    # Response is in pages -> while loop to request all pages
    while url:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        for message in data.get("value", []):
            messages_text.append(message_to_text(token, message))
        url = data.get("@odata.nextLink")
    return messages_text

def initialize_excel_sheet(excel_path, map_name):
    wb = load_workbook(excel_path)
    prefix_map = {
        "Postvak IN/1 Pieter Smit/Inv NL": "NL",
        "Postvak IN/1 Pieter Smit/Inv BE": "BE",
        "Postvak IN/1 Pieter Smit/Inv DE": "DE",
        "Postvak IN/1 Pieter Smit/Inv FR": "FR",
        "Postvak IN/1 Pieter Smit/Inv Nightliner": "Nightliner"
    }
    prefix = prefix_map.get(map_name, "")
    timestamp = datetime.now().strftime('%d-%m _ %H-%M')
    title = f"{prefix} - {timestamp}" if prefix else timestamp
    ws = wb.create_sheet(title=title)

    ws.cell(row=1, column=1).value = "Boekingsdatum"
    ws.cell(row=1, column=2).value = "Datum"
    ws.cell(row=1, column=3).value = "Tour"
    ws.cell(row=1, column=4).value = "Passagier"
    ws.cell(row=1, column=5).value = "Bestemming"
    ws.cell(row=1, column=6).value = "Prijs"
    ws.cell(row=1, column=7).value = "Fee"
    ws.cell(row=1, column=8).value = ""
    ws.cell(row=1, column=9).value = "Voorgesteld alternatief"
    ws.cell(row=1, column=10).value = "Missed/ Earned Savings"
    ws.cell(row=1, column=11).value = "Prijs Excel"
    ws.cell(row=1, column=12).value = "PNR"
    ws.cell(row=1, column=13).value = "Airline"

    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    row = 2
    ws.cell(row=row, column=1).value = "Vluchten"
    ws.cell(row=row, column=1).fill = highlight
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1).value = "Trein/bus"
    ws.cell(row=row, column=1).fill = highlight
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1).value = "Hotels"
    ws.cell(row=row, column=1).fill = highlight
    ws.cell(row=row, column=1).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1).value = "Refunds"
    ws.cell(row=row, column=1).fill = highlight
    ws.cell(row=row, column=1).font = Font(bold=True)

    wb.save(excel_path)
    return title

def build_prompt(email_text):
    return f"""Analyseer deze email (bijlagen starten met 'Text from PDF:') en extraheer reis-/boekingsgegevens volgens het schema.

Kies "type" op basis van de inhoud: "vlucht" of "trein/bus" voor reistickets, "hotel" voor accommodatie, "refund" voor terugbetalingen.

Veldbetekenis per type:

**Vlucht / trein-bus:**
- boekingsdatum: meestal de datum waarop de mail verzonden is (dd/mm/jjjj)
- datums: één entry per richting, eerst heenreis dan terugreis (indien aanwezig)
- bestemming: één entry per richting — "Stad van vertrek - Stad van aankomst"
- prijs: totale eindprijs
- PNR: boekingscode
- airline: naam van de maatschappij

**Hotel:**
- boekingsdatum: dd/mm/jjjj
- datum: "Incheck datum (dd/mm) - uitcheck datum (dd/mm)"
- bestemming: "Naam hotel, Stad"
- prijs: totale eindprijs
- PNR: boekingscode
- airline: naam van hotelmaatschappij

**Refund:**
- boekingsdatum, datum: dd/mm/jjjj
- bestemming: enkel de heenvlucht (bijv. "Brussel - Amsterdam")
- prijs: negatief bedrag (bijv. -123.45)
- PNR: boekingscode (mogelijk leeg bij refund)

Regels:
**Namen & Titels:**
- Verwijder titels: Mr, Mrs, Ms → niet in naam
- Format: "Voornaam Achternaam" (hoofdletters aan begin, NOOIT drukletters)
- LOT en TAP format "ACHTERNAAM VOORNAAM Mr" → draai om naar "Voornaam Achternaam"
- mytrip → niet naam van andere maatschappij zetten

**Plaatsnamen:**
- Altijd Nederlands en voluit (vertaal indien nodig)
- Alleen stadsnaam, geen luchthavennaam

**Vluchten:**
- Tussenstops samenvoegen: Amsterdam - Warschau - Wroclaw → Amsterdam - Wroclaw

**Maatschappijen:**
- Verkort: "LOT Airlines" → "LOT", "KLM Airlines" → "KLM", "TAP Air Portugal" → "TAP", "Expedia TAAP" → "Expedia", "Booking.com" → "Booking", Deutsche Bahn" → "DB"
- Altijd hoofdletter eerste letter
- NMBS: PNR = DNR van ticket

**Prijzen:**
- Euro: alleen cijfer (123.45)
- Andere valuta: cijfer + code (179.99 PLN)
-Nooit zelf een valuta omrekenen
- Expedia: neem bedrag bij "Subtotaal betaald aan Expedia of waar Mastercard achter staat", anders hoogste bedrag

**Hotels:**
- Passagier ontbreekt EN "Company" of "Pieter Smit" staat vermeld → passagier: "Company of Pieter Smit [BE/NL] [aantal]x"
- Naam bij Expedia: het bedrag waarbij staat "betaald aan Expedia" (dit is niet altijd het hoogste bedrag)

**Lege velden:**
- Leeg string "", NOOIT "N/A" of null

EMAIL:
\"\"\"
{email_text}
\"\"\"
"""


def extract_flight_data(email_text):
    response = client.messages.create(
        model="claude-sonnet-5",
        max_tokens=8000,
        output_config={"format": {"type": "json_schema","schema": EXTRACTION_SCHEMA}},
        messages=[{"role": "user", "content": build_prompt(email_text)}]
    )
    return next(b.text for b in response.content if b.type == "text")

def estimate_api_cost(mails):
    input_price_per_mtok = 3.00
    output_price_per_mtok = 15.00
    estimated_output_tokens_per_mail = 300  # ruwe schatting, JSON-antwoord is klein
    total_input_tokens = 0
    for m in mails:
        count = client.messages.count_tokens(
            model="claude-sonnet-5",
            output_config={
                "format": {
                    "type": "json_schema",
                    "schema": EXTRACTION_SCHEMA
                }
            },
            messages=[
                {"role": "user", "content": build_prompt(m)}
            ]
        )
        total_input_tokens += count.input_tokens

    input_cost = total_input_tokens / 1_000_000 * input_price_per_mtok
    estimated_output_tokens = len(mails) * estimated_output_tokens_per_mail
    output_cost = estimated_output_tokens / 1_000_000 * output_price_per_mtok

    return total_input_tokens, input_cost, output_cost

def extracted_flightdata_to_excel(ws, row, item):
    first_destination = True
    bestemmingen = item.get("bestemming",[])
    datums = item.get("datums", [])
    passagiers = item.get("passagiers", [])

    current_row = row

    for dest_index, dest in enumerate(bestemmingen):
        first_passenger = True

        datum = datums[dest_index].get("datum", "") if dest_index < len(datums) else ""

        for passenger in passagiers:

            ws.insert_rows(current_row)

            if first_passenger and first_destination:
                try:
                    date_obj = datetime.strptime(item.get("boekingsdatum", ""), "%d/%m/%Y").date()
                    cell = ws.cell(row=current_row, column=1, value=date_obj)
                    cell.number_format = "DD/MM/YYYY"
                except (ValueError, AttributeError):
                    ws.cell(row=current_row, column=1).value = item.get("boekingsdatum", "")

                try:
                    date_obj = datetime.strptime(datum, "%d/%m/%Y").date()
                    cell = ws.cell(row=current_row, column=2, value=date_obj)
                    cell.number_format = "DD/MM/YYYY"
                except (ValueError, AttributeError):
                    ws.cell(row=current_row, column=2).value = datum

                ws.cell(row=current_row, column=3).value = ""
                ws.cell(row=current_row, column=4).value = passenger.get("naam", "")
                ws.cell(row=current_row, column=5).value = dest.get("vlucht", "")
                ws.cell(row=current_row, column=6).value = ""

                try:
                    price = float(item["prijs"])
                    cell = ws.cell(row=current_row, column=11, value=price)
                    cell.number_format = "#,##0.00"
                except (ValueError, KeyError):
                    ws.cell(row=current_row, column=11).value = item.get("prijs", "")

                ws.cell(row=current_row, column=12).value = item.get("PNR", "")
                ws.cell(row=current_row, column=13).value = item.get("airline", "")

            else:
                try:
                    date_obj = datetime.strptime(datum, "%d/%m/%Y").date()
                    cell = ws.cell(row=current_row, column=2, value=date_obj)
                    cell.number_format = "DD/MM/YYYY"
                except (ValueError, KeyError):
                    ws.cell(row=current_row, column=2).value = datum

                ws.cell(row=current_row, column=4).value = passenger.get("naam", "")
                ws.cell(row=current_row, column=5).value = dest.get("vlucht", "")

            current_row += 1
            first_passenger = False

        first_destination = False

def extracted_data_to_excel(ws, row, item):
    passagiers = item.get("passagiers", [])
    first_passenger = True
    current_row = row

    for passenger in passagiers:
        ws.insert_rows(current_row)
        if first_passenger:
            try:
                date_obj = datetime.strptime(item["boekingsdatum"], "%d/%m/%Y").date()
                cell = ws.cell(row=current_row, column=1, value=date_obj)
                cell.number_format = "DD/MM/YYYY"
            except (ValueError, KeyError):
                ws.cell(row=current_row, column=1).value = item.get("boekingsdatum", "")
            try:
                date_obj = datetime.strptime(item["datum"], "%d/%m/%Y").date()
                cell = ws.cell(row=current_row, column=2, value=date_obj)
                cell.number_format = "DD/MM/YYYY"
            except (ValueError, KeyError):
                ws.cell(row=current_row, column=2).value = item.get("datum", "")
            ws.cell(row=current_row, column=3).value = ""
            ws.cell(row=current_row, column=5).value = item.get("bestemming", "")
            ws.cell(row=current_row, column=6).value = ""
            try:
                price = float(item["prijs"])
                cell = ws.cell(row=current_row, column=11, value=price)
                cell.number_format = "#,##0.00"
            except (ValueError, KeyError):
                ws.cell(row=current_row, column=11).value = item.get("prijs", "")
            ws.cell(row=current_row, column=12).value = item.get("PNR", "")
            ws.cell(row=current_row, column=13).value = item.get("airline", "")
        ws.cell(row=current_row, column=4).value = passenger.get("naam", "")
        current_row += 1
        first_passenger = False

def normalize_ticket_type(ticket_type):
    t = ticket_type.lower()
    if "vlucht" in t:
        return "vlucht"
    if "trein" in t or "bus" in t:
        return "trein/bus"
    if "hotel" in t:
        return "hotel"
    return t

def append_item_to_excel(item, excel_path, sheet_name):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    ticket_type = normalize_ticket_type(item.get("type", ""))

    section_names = {
        "vlucht": "Vluchten",
        "trein/bus": "Trein/bus",
        "hotel": "Hotels"
    }
    section_name = section_names.get(ticket_type, "Refunds")

    section_start = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == section_name:
            section_start = row
            break

    if section_start == 0:
        print(f"⚠️ WAARSCHUWING: Sectie '{section_name}' niet gevonden! Item overgeslagen.")
        return

    next_section_row = None
    for row in range(section_start + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value in ["Vluchten", "Trein/bus", "Hotels", "Refunds"]:
            next_section_row = row
            break

    if next_section_row:
        insert_row = next_section_row - 1
    else:
        insert_row = ws.max_row + 1

    if ticket_type == "vlucht" or ticket_type == "trein/bus":
        extracted_flightdata_to_excel(ws, insert_row, item)
    else:
        extracted_data_to_excel(ws, insert_row, item)

    format_excel_cells(ws)
    wb.save(excel_path)

def format_excel_cells(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = numbers.FORMAT_DATE_DDMMYY

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=11):
        for cell in row:
            cell.number_format = "#,##0.00"

# Starting mail extraction
def start(date_entry, map_var, excel_path, token, mailbox_lookup, progress_label):
    progress_label.configure(text="Begonnen met ophalen van emails", text_color="green")
    since_date_iso = date_entry.get_date().strftime("%Y-%m-%dT00:00:00Z")
    map_name = map_var.get()
    excel_file = excel_path.get()
    folder_id = mailbox_lookup.get(map_name)
    if not folder_id:
        print(f"Map '{map_name}' niet gevonden")
        progress_label.configure(text="Geen correcte map geselecteerd", text_color="red")
        return
    mails = get_all_email_data(token, folder_id, since_date_iso)
    amount_mails = len(mails)

    total_input_tokens, input_cost, estimated_output_cost = estimate_api_cost(mails)
    total_estimate = input_cost + estimated_output_cost

    print(amount_mails, " mails gevonden")
    progress_label.configure(text=f"{amount_mails} mails gevonden -- Verwacht kost = €{total_estimate}", text_color="green")

    # Excel blad maken
    try:
        sheet_name = initialize_excel_sheet(excel_file, map_name)
    except Exception as e:
        print(e)
        progress_label.configure(text="Excel sheet aanmaken mislukt (mogelijk staat Excel nog open)", text_color="red")
        return

    errors = 0
    number_of_handled_mails = 0
    for m in mails:
        try:
            json_string = extract_flight_data(m)
            print(json_string)
            parsed = json.loads(json_string)
            if isinstance(parsed, list):
                if len(parsed) > 0:
                    item = parsed[0]
                else:
                    print("⚠️ Lege lijst ontvangen van AI, email overgeslagen")
                    print("Mislukte mail: ", m)
                    errors += 1
                    continue
            else:
                item = parsed

            append_item_to_excel(item, excel_file, sheet_name)

            number_of_handled_mails += 1
            progress_label.configure(text=f"{number_of_handled_mails + errors} van de {amount_mails} verwerkt, {number_of_handled_mails} succesvol, {errors} mislukt", text_color="green")

        except json.JSONDecodeError as e:
            print("JSON kon niet gelezen worden:", e)
            print(json_string)
            errors += 1
            progress_label.configure(text=f"{number_of_handled_mails + errors} van de {amount_mails} verwerkt, {number_of_handled_mails} succesvol, {errors} mislukt", text_color="green")
            continue

        except Exception as e:
            print(f"Onverwachte error: {e}")
            errors += 1
            progress_label.configure(text=f"{number_of_handled_mails + errors} van de {amount_mails} verwerkt, {number_of_handled_mails} succesvol, {errors} mislukt", text_color="green")
            continue
    print(f'{number_of_handled_mails} van de {amount_mails} verwerkt en in Excel gezet')

def logout(app):
    if os.path.exists("token_cache.bin"):
        os.remove("token_cache.bin")
    app.destroy()
    sys.exit()

# Browsing files to select Excel file
def browse_file(excel_path, file_label, run_btn):
    filepath = filedialog.askopenfilename(
        title="Selecteer Excel-bestand",
        filetypes=[("Excel bestanden", "*.xlsx")],
    )
    if filepath:
        excel_path.set(filepath)
        file_label.configure(text=f"Geselecteerd: {os.path.basename(filepath)}", text_color="green")
        run_btn.configure(state=tkinter.NORMAL)

def get_mailboxes(token, parent_id=None, parent_path=""):
    headers = graph_headers(token)
    url = (f"{GRAPH_BASE}/me/mailFolders/{parent_id}/childFolders?$top=100" if parent_id else f"{GRAPH_BASE}/me/mailFolders?$top=100")
    folders = []
    while url:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        for folder in data.get("value", []):
            full_path = f"{parent_path}/{folder['displayName']}" if parent_path else folder["displayName"]
            folders.append({"path": full_path, "id": folder["id"]})
            if folder.get("childFolderCount", 0) > 0:
                folders.extend(get_mailboxes(token, folder["id"], full_path))
        url = data.get("@odata.nextLink")
    return folders

def start_main_thread(date_entry, map_var, excel_path, token, mailbox_lookup, progress_label):
    thread = threading.Thread(target=lambda: start(date_entry, map_var, excel_path, token, mailbox_lookup, progress_label))
    thread.start()

def build_ui_content(app, mailboxes, token):
    mailbox_paths = [m["path"] for m in mailboxes]
    mailbox_lookup = {m["path"]: m["id"] for m in mailboxes}

    # UI Elements
    date_title = customtkinter.CTkLabel(app, text="Vul de datum van de eerste mail in")
    date_title.pack(pady=(10,0))

    date_entry = DateEntry(app, date_pattern='dd/mm/yyyy')
    date_entry.configure(font=tkFont.Font(size=14))
    date_entry.pack(padx=10, pady=(10,20))

    map_title = customtkinter.CTkLabel(app, text="Vul de naam van de map in (vb. NL)")
    map_title.pack()

    map_var = tkinter.StringVar()
    cb = ttk.Combobox(app, values=mailbox_paths, textvariable=map_var, width=25, height=40)
    cb.configure(font=tkFont.Font(size=14) )
    cb.pack(pady=(0,10))

    def filter_mailboxes(event):
        typed = map_var.get().lower()
        if typed == "":
            cb["values"] = mailbox_paths
        else:
            cb["values"] = [m for m in mailbox_paths if typed in m.lower()]

    cb.bind("<KeyRelease>", filter_mailboxes)

    excel_path = tkinter.StringVar()
    file_label = customtkinter.CTkLabel(app, text="Geen bestand geselecteerd", text_color="red")
    progress_label = customtkinter.CTkLabel(app, text = "")

    run_btn = customtkinter.CTkButton(app, text="Start", command=lambda: start_main_thread(date_entry, map_var, excel_path, token, mailbox_lookup, progress_label))
    browse_btn = customtkinter.CTkButton(app, text="Kies Excel-bestand", command=lambda: browse_file(excel_path, file_label, run_btn))
    browse_btn.pack(pady=(20,0))
    file_label.pack()

    run_btn.pack(pady=(20,10))
    run_btn.configure(state=tkinter.DISABLED)

    logout_btn = customtkinter.CTkButton(app, text="Log uit", command=lambda: logout(app))
    logout_btn.pack(pady=(20,10))

    finish_label = customtkinter.CTkLabel(app, text = "")
    finish_label.pack()

    progress_label.pack()

    return app

def finish_login(app, status_label, mailboxes, token):
    status_label.destroy()
    build_ui_content(app, mailboxes, token)

def main():
    # System setting
    customtkinter.set_appearance_mode("System")
    customtkinter.set_default_color_theme("blue")
    # App frame
    app = customtkinter.CTk()
    app.geometry("720x480")
    app.title("Excel assistent")

    status_label = customtkinter.CTkLabel(app, text="Bezig met inloggen...")
    status_label.pack(pady=20)

    def device_flow_callback(flow):
        app.after(0, lambda: status_label.configure(
            text=f"Ga naar {flow['verification_uri']} en voer code {flow['user_code']} in."
        ))
        webbrowser.open(flow["verification_uri"])

    def login_worker():
        try:
            token = get_token(device_flow_callback)
            mailboxes = get_mailboxes(token)
            app.after(0, lambda: finish_login(app, status_label, mailboxes, token))
        except Exception as e:
            app.after(0, lambda: status_label.configure(text=f"Login mislukt: {e}", text_color="red"))

    threading.Thread(target=login_worker, daemon=True).start()
    app.mainloop()

if __name__ == "__main__":
    main()
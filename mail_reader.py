import imaplib
import email
import os
import tempfile

from bokeh.util.terminal import white
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from PyPDF2 import PdfReader
from openai import OpenAI
import json
from openpyxl import load_workbook
from dotenv import load_dotenv
import tkinter
import customtkinter
import threading
from tkinter import filedialog
from tkinter import ttk
import tkinter.font as tkFont

load_dotenv()

usr = os.getenv("EMAIL_USER")
pwd = os.getenv("EMAIL_PASS")

client = OpenAI(api_key=os.getenv("openai_api_key"))

def clean(text):
    # clean text for creating a folder
    return "".join(c if c.isalnum() else "_" for c in text)

def email_to_text(mail, date, map):
    map = '"' + map + '"'
    mail.select(map)
    try:
        typ, data = mail.search(None, "SINCE", date)
    except:
        print("Wrong date format")
        finish_label.configure(text="Ongeldige datum of map", text_color="red")
    extracted_mails = []
    for i in sorted(data[0].split(), key = int, reverse=True):
        complete_mail = ""
        res, msg = mail.fetch(i, "(RFC822)")
        for response in msg:
            if isinstance(response, tuple):
                # parse a bytes email into a message object
                msg = email.message_from_bytes(response[1])
                if msg.is_multipart():
                    # iterate over email parts
                    for part in msg.walk():
                        # extract content type of email
                        content_type = part.get_content_type()
                        content_disposition = str(part.get("Content-Disposition"))
                        try:
                            # get the email body
                            body = part.get_payload(decode=True).decode()
                        except:
                            pass
                        if content_type == "text/plain" and "attachment" not in content_disposition:
                            # print text/plain emails and skip attachments
                            complete_mail += body
                        elif "attachment" in content_disposition:
                            # download attachment
                            filename = part.get_filename()
                            if filename and filename.lower().endswith(".pdf"):
                                # only handle if file is PDF
                                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                                    tmp_file.write(part.get_payload(decode=True))
                                    tmp_filepath = tmp_file.name

                                try:
                                    reader = PdfReader(tmp_filepath)
                                    pdf_text = "\n".join([page.extract_text() for page in reader.pages])
                                    complete_mail += "Text from PDF:"
                                    complete_mail += pdf_text
                                except Exception as e:
                                    print(f"Error while reading PDF {filename}: {e}")
                                finally:
                                    os.remove(tmp_filepath)
                else:
                    # extract content type of email
                    content_type = msg.get_content_type()
                    # get the email body
                    body = msg.get_payload(decode=True).decode()
                    if content_type == "text/plain":
                        # print only text email parts
                        complete_mail += body
        extracted_mails += [complete_mail]

    # close the connection and logout
    mail.close()
    mail.logout()
    return extracted_mails

def extract_flight_data(email_text):
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "Je bent een behulpzame assistent die e-mails omzet in gestructureerde data."},
            {"role": "user", "content": f"""
Je bent een slimme data-extractie assistent voor een reisagent. Je krijgt telkens een email en mogelijke de toegevoegde bijlagen, de bijlagen start altijd met 'Text from PDF:'.
Haal uit de mail de vluchtdetails en geef ze als JSON terug in dit formaat voor vluvhten, treinen, bussen en transfers:
[
    {{
        "type": "Opties tussen: vlucht of trein/bus",
        "boekingsdatum": "De datum waarop de boeking vastgelegd is (de dag van betaling of meestal waarop de mail gestuurd is) (mm/dd/jjjj)",
        "datum": "De datum van de reis (mm/dd/jjjj)",
        "passagier": "Volledige naam",
        "bestemming": "Van - Naar",
        "prijs": "De totale eindprijs op het ticket, staat meestal achter total amount of paid by card of iets gelijkaardigs (vb 123.45)",
        "PNR": "De boekingscode van op het ticket",
        "airline": "De naam van de maatschappij waarbij het ticket geboekt is"
    }},
]
Deze vorm voor hotels:
[
    {{
        "type": "hotel",
        "boekingsdatum": "De datum waarop de boeking vastgelegd is (de dag van betaling, meestal de dag waarop de mail gestuurd is) (mm/dd/jjjj)",
        "datum": "Incheck datum (mm/dd/jjjj)",
        "passagier": "Volledige naam",
        "bestemming": "Naam hotel, Stad",
        "prijs": "De totale eindprijs op het ticket, staat meestal achter total amount of paid by card of iets gelijkaardigs (vb 123.45)",
        "PNR": "De boekingscode van op het ticket",
        "airline": "De naam van de maatschappij waarbij het ticket geboekt is"
    }},
]
Deze vorm voor refunds:
[
    {{
        "type": "refund",
        "boekingsdatum": "De datum waarop de boeking vastgelegd is (de dag van betaling of meestal waarop de mail gestuurd is) (mm/dd/jjjj)",
        "datum": "Incheck datum (mm/dd/jjjj)",
        "passagier": "Volledige naam",
        "bestemming": "Naam hotel, Stad",
        "prijs": "De totale eindprijs op het ticket, staat meestal achter total amount of paid by card of iets gelijkaardigs (vb 123.45)",
        "PNR": "De boekingscode van op het ticket",
        "airline": "De naam van de maatschappij waarbij het ticket geboekt is"
    }},
]
Antwoord enkel met geldige JSON. Geen commentaar, geen uitleg, geen codeblok-tekens.
Als er meerdere namen op het ticket staan moet elke passagier een eigen object zijn met alle info, maar enkel bij de eerste passagier mag de totale prijs staan en bij de andere moet de prijs leeg zijn.
Wanneer er bij een hotel enkel Company of Pieter Smit staat en geen passagiersnaam wordt passagier bij hotel: "Company of Pieter Smit 'BE/NL als dit vermeld is' '#aantal kamers' x
Als er een heen en terugvlucht op het ticket staat moeten beide een object zijn, maar de prijs mag enkel ingevuld zijn bij de heenvlucht.
Wanneer er een overstap gemaakt wordt hoeft dit niet vermeld te worden, dus bijvoorbeeld Amsterdam - Warschau en Warschau - Wroclaw op dezelfde datum wordt Amsterdam - Wroclaw.
Een transfer hoort bij trein/ bus
Als er Mr of iets gelijkaardigs in de naam zit staat dit voor meneer en moet je dit niet mee in de naam zetten
Plaatsnamen moeten in het nederlands en altijd voluit
Wees consisten in de namen bijvoorbeeld: LOT of LOT airlines is altijd LOT, KLM airlines is gewoon KLM
Namen van passagiers beginnen met een hoofdletter, maar mogen nooit in drukletters
Namen van maatschappijen altijd met hoofdletter
Sommige mails bevatten ook pdf's, deze zijn meegeplakt in de tekst van de mail en het kan zijn dat info van een vlucht dubbel is. Zorg dus dat je nooit twee objecten waarbij naam, bestemmnig, datum en boekingsnummer bij alle twee hetzelfde zijn.
EXTREEM BELANGRIJK: IK WIL BIJ DE PRIJS NOOIT EUR ZIEN STAAN, EEN PRIJS IN EURO IS GEWOON EEN GETAL. EEN ANDERE VALUTA DAN EUR MOET WEL VERMELD WORDEN (VB/ 179.99 PLN)
EXTREEM BELANGRIJK: IK WIL NOOIT ERGENS N/A ZIEN STAAN, DIT MOET HANDMATIG VERWIJDERD WORDEN EN HET VAKJE KAN DUS BETER ONMIDDELIJK OPENGELATEN WORDEN

EMAIL:
\"\"\"
{email_text}
\"\"\"
"""}
        ],
        temperature=0.0
    )
    print(response.choices[0].message.content)
    return response.choices[0].message.content

def extracted_data_to_excel(ws, data):
    row = ws.max_row + 1
    for item in data:
        ws.cell(row=row, column=1).value = item.get("boekingsdatum", "")
        ws.cell(row=row, column=2).value = item.get("datum", "")
        ws.cell(row=row, column=3).value = ""
        ws.cell(row=row, column=4).value = item.get("passagier", "")
        ws.cell(row=row, column=5).value = item.get("bestemming", "")
        ws.cell(row=row, column=6).value = ""
        ws.cell(row=row, column=7).value = item.get("prijs", "")
        ws.cell(row=row, column=8).value = item.get("PNR", "")
        ws.cell(row=row, column=9).value = item.get("airline", "")
        row += 1
    return row

def write_json_to_excel(data, excel_path):
    data_without_duplicates = []
    i = 0
    while i < len(data):
        j = 0
        duplicate = False
        while j < i:
            if data[i] == data[j]:
                duplicate = True
            j += 1
        if not duplicate:
            data_without_duplicates += [data[i]]
        i += 1

    flights = []
    trains = []
    hotels = []
    refunds = []
    for item in data_without_duplicates:
        ticket_type = item.get("type", "")
        if ticket_type == "vlucht":
            flights += [item]
        elif ticket_type == "trein/bus":
            trains += [item]
        elif ticket_type == "hotel":
            hotels += [item]
        else:
            refunds += [item]

    wb = load_workbook(excel_path)
    ws = wb.create_sheet(title=f"{datetime.now().strftime('%Y-%m-%d')}")

    ws.cell(row=1, column=1).value = "Boekingsdatum"
    ws.cell(row=1, column=2).value = "Datum"
    ws.cell(row=1, column=3).value = "Tour"
    ws.cell(row=1, column=4).value = "Passagier"
    ws.cell(row=1, column=5).value = "Bestemming"
    ws.cell(row=1, column=6).value = "Prijs"
    ws.cell(row=1, column=7).value = "Prijs Excel"
    ws.cell(row=1, column=8).value = "PNR"
    ws.cell(row=1, column=9).value = "Airline"

    row = 2
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    if len(flights) != 0:
        ws.cell(row=row, column=1).value = "Vluchten"
        ws.cell(row=row, column=1).fill = highlight
        ws.cell(row=row, column=1).font = Font(bold=True)
        row = extracted_data_to_excel(ws, flights)

    if len(trains) != 0:
        ws.cell(row=row + 1, column=1).value = "Trein/bus"
        ws.cell(row=row + 1, column=1).fill = highlight
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        row = extracted_data_to_excel(ws, trains)

    if len(hotels) != 0:
        ws.cell(row=row + 1, column=1).value = "Hotels"
        ws.cell(row=row + 1, column=1).fill = highlight
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        row = extracted_data_to_excel(ws, hotels)

    if len(refunds) != 0:
        ws.cell(row=row + 1, column=1).value = "Refunds"
        ws.cell(row=row + 1, column=1).fill = highlight
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        extracted_data_to_excel(ws, refunds)

    wb.save(excel_path)
    print(f"✅ Gegevens toegevoegd aan {excel_path}")

def main(mail, date, map, excel):
    finish_label.configure(text="Bezig met mails lezen en in Excel zetten!", text_color="white")
    progress_label.configure(text= "")
    email_list = email_to_text(mail, date, map)
    email_full_text = ""
    json_items = []
    number_of_mails = len(email_list)
    number_of_handled_mails = 0
    progress_label.configure(text="0 van de " + str(number_of_mails) + " uitgelezen")
    for m in email_list:
        email_full_text += m
        json_string = extract_flight_data(m)
        try:
            item = json.loads(json_string)
        except json.JSONDecodeError as e:
            finish_label.configure(text="AI geeft onleesbare vorm terug", text_color="red")
            print("JSON kon niet gelezen worden:", e)
            print(json_string)
            quit()
        json_items += item
        number_of_handled_mails += 1
        progress_label.configure(text= str(number_of_handled_mails) + " van de " + str(number_of_mails) + " uitgelezen")

    write_json_to_excel(json_items, excel)
    finish_label.configure(text="Klaar! Selecteer een andere map of sluit het programma.")

def start_main_thread():
    thread = threading.Thread(target=lambda: main(mail, date.get(), map.get(), excel_path.get()))
    thread.start()

def browse_file():
    filepath = filedialog.askopenfilename(
        title="Selecteer Excel-bestand",
        filetypes=[("Excel bestanden", "*.xlsx")],
    )
    if filepath:
        excel_path.set(filepath)
        file_label.configure(text=f"Geselecteerd: {os.path.basename(filepath)}", text_color="green")
        run_btn.configure(state=tkinter.NORMAL)


mail = imaplib.IMAP4_SSL("imap.gmail.com", 993)
mail.login(usr, pwd)
status, mailboxes = mail.list()
mailbox_names = []
if status == 'OK':
    for mbox in mailboxes:
        decoded = mbox.decode()
        parts = decoded.split(' "/" ')
        if len(parts) == 2:
            name = parts[1].strip('"')
            mailbox_names += [str(name)]
else:
    print("Kon mailboxen niet ophalen")

# System setting
customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")

# App frame
app = customtkinter.CTk()
app.geometry("720x480")
app.title("Excel assistent")

# UI Elements
date_title = customtkinter.CTkLabel(app, text="Vul de datum van de eerste mail in (vb. 01-Aug-2025)")
date_title.pack(pady=(10,0))

date = tkinter.StringVar()
date_input = customtkinter.CTkEntry(app, width=100, height=40, textvariable=date)
date_input.pack(pady=(0,10))

map_title = customtkinter.CTkLabel(app, text="Vul de naam van de map in (vb. NL)")
map_title.pack()

map = tkinter.StringVar()
cb = ttk.Combobox(app, values=mailbox_names, textvariable=map, width=50, height=40)
cb.configure(font=tkFont.Font(size=14) )
cb.pack(pady=(0,10))

excel_path = tkinter.StringVar()
browse_btn = customtkinter.CTkButton(app, text="Kies Excel-bestand", command=browse_file)
browse_btn.pack(pady=(10,0))

file_label = customtkinter.CTkLabel(app, text="Geen bestand geselecteerd", text_color="red")
file_label.pack()


finish_label = customtkinter.CTkLabel(app, text = "")
finish_label.pack()

progress_label = customtkinter.CTkLabel(app, text = "")
progress_label.pack()

run_btn = customtkinter.CTkButton(app, text="Start", command=start_main_thread)
run_btn.pack()
run_btn.configure(state=tkinter.DISABLED)

# Run app
app.mainloop()


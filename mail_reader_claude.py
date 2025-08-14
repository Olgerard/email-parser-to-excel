import base64
import os
import sys
import tempfile
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from PyPDF2 import PdfReader
from anthropic import Anthropic
import json
from openpyxl import load_workbook
from dotenv import load_dotenv
import tkinter
import customtkinter
import threading
from tkinter import filedialog
from tkinter import ttk
import tkinter.font as tkFont
from tkcalendar import DateEntry
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

load_dotenv()

client = Anthropic(api_key=os.getenv("claude_api_key"))

if not client:
    print("❌ Vul eerst je .env bestand in!")
    quit()

def email_to_text_gmail(service, date, label):
    query = f'after:{datetime.strptime(date, "%d-%b-%Y").strftime("%Y/%m/%d")}'
    if label:
        all_labels = service.users().labels().list(userId='me').execute().get('labels', [])
        label_id = next((l['id'] for l in all_labels if l['name'] == label), None)
    else:
        label_id = None
        finish_label.configure(text="Geen map ingevuld!", text_color="red")
        quit()

    results = service.users().messages().list(
        userId='me',
        q=query,
        labelIds=[label_id] if label_id else None
    ).execute()

    messages = results.get('messages', [])
    messages.reverse()
    extracted_mails = []

    for msg in messages:
        m = service.users().messages().get(userId='me', id=msg['id']).execute()
        payload = m['payload']
        parts = payload.get('parts', [])
        complete_mail = ""

        def walk_parts(parts_list):
            nonlocal complete_mail
            for part in parts_list:
                mime_type = part.get('mimeType', '')
                body_data = part.get('body', {}).get('data')
                if body_data:
                    decoded = base64.urlsafe_b64decode(body_data).decode(errors="ignore")
                    if mime_type == 'text/plain':
                        complete_mail += decoded
                    elif mime_type == 'application/pdf':
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                            tmp_file.write(base64.urlsafe_b64decode(body_data))
                            tmp_filepath = tmp_file.name
                        try:
                            reader = PdfReader(tmp_filepath)
                            pdf_text = "\n".join([page.extract_text() for page in reader.pages])
                            complete_mail += "Text from PDF:" + pdf_text
                        finally:
                            os.remove(tmp_filepath)
                if 'parts' in part:
                    walk_parts(part['parts'])

        walk_parts(parts)
        extracted_mails.append(complete_mail)

    return extracted_mails

def extract_flight_data(email_text):
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=20000,
        temperature=1,
        messages=[
            {"role": "user", "content": f"""
Je bent een slimme data-extractie assistent voor een reisagent. Je krijgt telkens een email en mogelijke de toegevoegde bijlagen, de bijlagen start altijd met 'Text from PDF:'.
Haal uit de mail de vluchtdetails en geef ze als JSON terug in dit formaat:
[
    {{
        "type": "Opties tussen: vlucht of trein/bus",
        "boekingsdatum": In 95% van de gevallen de datum waarop de mail gestuurd is (mm/dd/jjjj)",
        "datum": "De datum van de reis (mm/dd/jjjj)",
        "passagier": "Volledige naam",
        "bestemming": "Stad van vertrek - Stad van aankomst",
        "prijs": "De totale eindprijs op het ticket, staat meestal achter total amount of paid by card of iets gelijkaardigs (vb 123.45)",
        "PNR": "De boekingscode van op het ticket",
        "airline": "De naam van de maatschappij waarbij het ticket geboekt is"
    }},
]
Deze vorm voor hotels:
[
    {{
        "type": "hotel",
        "boekingsdatum": "In 95% van de gevallen de datum waarop de mail gestuurd is (mm/dd/jjjj) (mm/dd/jjjj)",
        "datum": "Incheck datum (mm/dd/jjjj)",
        "passagier": "Volledige naam",
        "bestemming": "Naam hotel, Stad",
        "prijs": "De totale eindprijs op het ticket, staat meestal achter total amount of paid by card of iets gelijkaardigs (vb 123.45)",
        "PNR": "De boekingscode van op het ticket",
        "airline": "De naam van de maatschappij waarbij het ticket geboekt is"
    }},
]
Deze vorm voor refunds:
[
    {{
        "type": "refund",
        "boekingsdatum": "In 95% van de gevallen de datum waarop de mail gestuurd is (mm/dd/jjjj) (mm/dd/jjjj)",
        "datum": "Incheck datum (mm/dd/jjjj)",
        "passagier": "Volledige naam",
        "bestemming": "Naam hotel, Stad",
        "prijs": "De totale eindprijs op het ticket met een minteken (vb -123.45)",
        "PNR": "De boekingscode van op het ticket",
        "airline": "De naam van de maatschappij waarbij het ticket geboekt is"
    }},
]
De informatie in de JSON objecten wordt in Excel gezet. Houdt hier rekening mee (bv datums: niet 08/20/2025, maar 8/20/2025). Blijf dus ook constistent en behoudt de amerikaane vorm voor datums en getallen
Als er meerdere namen op het ticket staan moet elke passagier een eigen object zijn met alle info, maar enkel bij de eerste passagier mag de totale prijs staan en bij de andere moet de prijs leeg zijn.
Als er een heen en terugvlucht op het ticket staat moeten beide een object zijn, maar de prijs mag enkel ingevuld zijn bij de heenvlucht.
Als er meerdere vluchten en meerdere passagiers in 1 mail staan dan doen alle passagiers al de vluchten. Zet dan eerst voor alle namen de eerste vlucht en dan voor alle namen de andere vlucht.
Wanneer er bij een hotel enkel Company of Pieter Smit staat en geen passagiersnaam wordt passagier bij hotel: "Company of Pieter Smit 'BE/NL als dit vermeld is' '#aantal kamers' x.
Wanneer er een overstap gemaakt wordt hoeft dit niet vermeld te worden, dus bijvoorbeeld Amsterdam - Warschau en Warschau - Wroclaw op dezelfde datum wordt Amsterdam - Wroclaw.
Een transfer hoort bij trein/ bus.
Als er mr of iets gelijkaardigs in de naam zit staat dit voor meneer en moet je dit niet mee in de naam zetten.
Plaatsnamen moeten altijd in het Nederlands (als het op het ticket niet het geval is moet je zelf vertalen) en altijd voluit. Enkel de naam van de stad is nodig de naam van de luchthaven hoeft niet.
Wees consisten in de namen bijvoorbeeld: LOT of LOT airlines is altijd LOT, KLM airlines is gewoon KLM, Expedia en niet Expedia TAAP, NMBS, TAP Air Portugal is TAP... .
Namen van passagiers beginnen met een hoofdletter, maar mogen nooit in drukletters.
Namen van maatschappijen altijd met hoofdletter.
Er kan nooit een negatief bedrag zijn wanneer het type niet refund is.
Bij LOT staan de namen in deze vorm: NAAM VOORNAAM Mr let er op dat je dit omdraait om consistent met de rest te blijven.
Brussels Charleroi = Charleroi
Bij refunds van KLM ga je enkel het boekingsnummer, boekingsdatum (datum van mail) en misschien de naam terugvinden, laat de rest dus gewoon open
EXTREEM BELANGRIJK: IK WIL BIJ DE PRIJS NOOIT EUR ZIEN STAAN, EEN PRIJS IN EURO IS GEWOON EEN GETAL. EEN ANDERE VALUTA DAN EUR MOET WEL VERMELD WORDEN (VB/ 179.99 PLN).
EXTREEM BELANGRIJK: IK WIL NOOIT ERGENS N/A ZIEN STAAN, DIT MOET HANDMATIG VERWIJDERD WORDEN EN HET VAKJE KAN DUS BETER ONMIDDELIJK OPENGELATEN WORDEN.
EXTREEM BELANGRIJK: Geef alleen geldige JSON terug, zonder enige toelichting of tekst errond zoals ```json. Je respons moet dus exact in de vorm van de voorbeelden staan en mag hier absoluut niet van afwijken.
EMAIL:
\"\"\"
{email_text}
\"\"\"
"""}
        ]
    )
    return response.content[0].text

def extracted_data_to_excel(ws, data):
    row = ws.max_row + 1
    for item in data:
        try:
            date_obj = datetime.strptime(item["boekingsdatum"], "%m/%d/%Y").date()
            cell = ws.cell(row=row, column=1, value=date_obj)
            cell.number_format = "MM/DD/YYYY"
        except ValueError:
            ws.cell(row=row, column=1).value = item.get("boekingsdatum", "")
        try:
            date_obj = datetime.strptime(item["datum"], "%m/%d/%Y").date()
            cell = ws.cell(row=row, column=2, value=date_obj)
            cell.number_format = "MM/DD/YYYY"
        except ValueError:
            ws.cell(row=row, column=2).value = item.get("datum", "")
        ws.cell(row=row, column=3).value = ""
        ws.cell(row=row, column=4).value = item.get("passagier", "")
        ws.cell(row=row, column=5).value = item.get("bestemming", "")
        ws.cell(row=row, column=6).value = ""
        try:
            price = float(item["prijs"])
            cell = ws.cell(row=row, column=7, value=price)
            cell.number_format = "#,##0.00"
        except ValueError:
            ws.cell(row=row, column=7).value = item.get("prijs", "")
        ws.cell(row=row, column=8).value = item.get("PNR", "")
        ws.cell(row=row, column=9).value = item.get("airline", "")
        row += 1
    return row

def write_json_to_excel(data, excel_path, map):
    data_without_duplicates = []
    i = 0
    while i < len(data):
        j = 0
        duplicate = False
        while j < i:
            if data[i] == data[j]:
                duplicate = True
            j += 1
        if not duplicate:
            data_without_duplicates += [data[i]]
        i += 1

    flights = []
    trains = []
    hotels = []
    refunds = []
    for item in data_without_duplicates:
        ticket_type = item.get("type", "")
        if ticket_type == "vlucht":
            flights += [item]
        elif ticket_type == "trein/bus":
            trains += [item]
        elif ticket_type == "hotel":
            hotels += [item]
        else:
            refunds += [item]

    wb = load_workbook(excel_path)
    ws = wb.create_sheet(title=f"{datetime.now().strftime('%m-%d _ %H-%M')}")

    ws.cell(row=1, column=1).value = "Boekingsdatum"
    ws.cell(row=1, column=2).value = "Datum"
    ws.cell(row=1, column=3).value = "Tour"
    ws.cell(row=1, column=4).value = "Passagier"
    ws.cell(row=1, column=5).value = "Bestemming"
    ws.cell(row=1, column=6).value = "Prijs"
    ws.cell(row=1, column=7).value = "Prijs Excel"
    ws.cell(row=1, column=8).value = "PNR"
    ws.cell(row=1, column=9).value = "Airline"

    row = 2
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    if len(flights) != 0:
        ws.cell(row=row, column=1).value = "Vluchten"
        ws.cell(row=row, column=1).fill = highlight
        ws.cell(row=row, column=1).font = Font(bold=True)
        row = extracted_data_to_excel(ws, flights)

    if len(trains) != 0:
        ws.cell(row=row + 1, column=1).value = "Trein/bus"
        ws.cell(row=row + 1, column=1).fill = highlight
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        row = extracted_data_to_excel(ws, trains)

    if len(hotels) != 0:
        ws.cell(row=row + 1, column=1).value = "Hotels"
        ws.cell(row=row + 1, column=1).fill = highlight
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        row = extracted_data_to_excel(ws, hotels)

    if len(refunds) != 0:
        ws.cell(row=row + 1, column=1).value = "Refunds"
        ws.cell(row=row + 1, column=1).fill = highlight
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        extracted_data_to_excel(ws, refunds)

    wb.save(excel_path)
    print(f"✅ Gegevens toegevoegd aan {excel_path}")

def main(service, date, map, excel):
    finish_label.configure(text="Bezig met mails lezen en in Excel zetten!", text_color="white")
    progress_label.configure(text= "")
    email_list = email_to_text_gmail(service, date, map)
    json_items = []
    number_of_mails = len(email_list)
    number_of_handled_mails = 0
    progress_label.configure(text="0 van de " + str(number_of_mails) + " uitgelezen")
    for m in email_list:
        json_string = extract_flight_data(m)
        print(json_string)
        try:
            item = json.loads(json_string)
            if not isinstance(item, list):
                print("⚠️ Verwacht list, kreeg:", type(item), item, " mail nummer: " ,number_of_handled_mails + 1)
                continue
        except json.JSONDecodeError as e:
            finish_label.configure(text="AI geeft onleesbare vorm terug", text_color="red")
            print("JSON kon niet gelezen worden:", e)
            print(json_string)
            quit()
        json_items += item
        number_of_handled_mails += 1
        progress_label.configure(text= str(number_of_handled_mails) + " van de " + str(number_of_mails) + " uitgelezen")
    for i, item in enumerate(json_items):
        if not isinstance(item, dict):
            print(f"Fout item op index {i}: {item} (type {type(item)})")
    write_json_to_excel(json_items, excel, map)
    finish_label.configure(text="Klaar! Selecteer een andere map of sluit het programma.", text_color="green")

def start_main_thread():
    thread = threading.Thread(target=lambda: main(service, date_entry.get_date().strftime("%d-%b-%Y"), map.get(), excel_path.get()))
    thread.start()

def browse_file():
    filepath = filedialog.askopenfilename(
        title="Selecteer Excel-bestand",
        filetypes=[("Excel bestanden", "*.xlsx")],
    )
    if filepath:
        excel_path.set(filepath)
        file_label.configure(text=f"Geselecteerd: {os.path.basename(filepath)}", text_color="green")
        run_btn.configure(state=tkinter.NORMAL)


def get_gmail_service():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)

def logout():
    if os.path.exists("token.json"):
        os.remove("token.json")
        app.destroy()
        sys.exit()
    print("Uitgelogd: token verwijderd.")

service = get_gmail_service()

# labels ophalen
results = service.users().labels().list(userId='me').execute()
labels = results.get('labels', [])
mailbox_names = [label['name'] for label in labels]

# System setting
customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")

# App frame
app = customtkinter.CTk()
app.geometry("720x480")
app.title("Excel assistent")

# UI Elements
date_title = customtkinter.CTkLabel(app, text="Vul de datum van de eerste mail in (vb. 01-Aug-2025)")
date_title.pack(pady=(10,0))

date = tkinter.StringVar()
date_entry = DateEntry(app, date_pattern='dd/mm/yyyy')
date_entry.configure(font=tkFont.Font(size=14))
date_entry.pack(padx=10, pady=(10,20))

map_title = customtkinter.CTkLabel(app, text="Vul de naam van de map in (vb. NL)")
map_title.pack()

map = tkinter.StringVar()
cb = ttk.Combobox(app, values=mailbox_names, textvariable=map, width=25, height=40)
cb.configure(font=tkFont.Font(size=14) )
cb.pack(pady=(0,10))

excel_path = tkinter.StringVar()
browse_btn = customtkinter.CTkButton(app, text="Kies Excel-bestand", command=browse_file)
browse_btn.pack(pady=(20,0))

file_label = customtkinter.CTkLabel(app, text="Geen bestand geselecteerd", text_color="red")
file_label.pack()

run_btn = customtkinter.CTkButton(app, text="Start", command=start_main_thread)
run_btn.pack(pady=(20,10))
run_btn.configure(state=tkinter.DISABLED)

logout_btn = customtkinter.CTkButton(app, text="Log uit", command=logout)
logout_btn.pack(pady=(20,10))

finish_label = customtkinter.CTkLabel(app, text = "")
finish_label.pack()

progress_label = customtkinter.CTkLabel(app, text = "")
progress_label.pack()


# Run app
app.mainloop()